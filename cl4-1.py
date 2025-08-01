import os
import csv
import itertools
import sys
import re
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")
    print("Will create CSV file instead of Excel file with colors.")

def xor(a, b):
    """XOR operation"""
    return a != b

def parse_logical_expression(expr, variables):
    """Parse a logical expression and return a function to evaluate it"""
    # Clean the expression
    expr = expr.strip()
    original_expr = expr
    
    # Handle complex expressions with multiple operators more carefully
    # First, handle XOR (⊕) in parentheses and standalone
    expr = re.sub(r'([A-Z])\s*⊕\s*([A-Z])', r'xor(\1, \2)', expr)
    
    # Handle negation ¬ - be more careful with parentheses and variables
    expr = re.sub(r'¬\(([^)]+)\)', r'(not (\1))', expr)
    expr = re.sub(r'¬([A-Z])', r'(not \1)', expr)
    
    # Handle biconditional (↔) - more robust parsing
    while '↔' in expr:
        # Find the biconditional and split it properly
        parts = expr.split('↔', 1)  # Split only on first occurrence
        if len(parts) == 2:
            left = parts[0].strip()
            right = parts[1].strip()
            expr = f'(({left}) == ({right}))'
        else:
            break
    
    # Handle implication (→) - more robust parsing
    while '→' in expr:
        parts = expr.split('→', 1)  # Split only on first occurrence
        if len(parts) == 2:
            left = parts[0].strip()
            right = parts[1].strip()
            expr = f'((not ({left})) or ({right}))'
        else:
            break
    
    # Handle logical operators
    expr = expr.replace('∧', ' and ')
    expr = expr.replace('∨', ' or ')
    
    print(f"Original: {original_expr}")
    print(f"Parsed: {expr}")
    
    # Create a function to evaluate the expression
    def evaluate(**kwargs):
        # Create local variables for evaluation
        local_vars = {}
        for var in variables:
            local_vars[var] = bool(kwargs.get(var, False))
        local_vars['xor'] = xor
        
        try:
            result = eval(expr, {"__builtins__": {}}, local_vars)
            return bool(result)
        except Exception as e:
            print(f"Error evaluating expression: {expr}")
            print(f"Variables: {local_vars}")
            print(f"Error details: {e}")
            return False
    
    return evaluate

def read_input_file(filename):
    """Read and parse the input file"""
    premises = []
    conclusion = None
    variables = set()
    
    with open(filename, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    # Process each line
    for line in lines:
        line = line.strip()
        # Remove quotes if present
        if line.startswith('"') and line.endswith('"'):
            line = line[1:-1]
        
        # Skip empty lines and certain keywords
        if not line or line.startswith('Variables:') or line == 'Premises:':
            continue
            
            # Find premises (numbered lines, including those starting with #)
        premise_match = re.match(r'#*(\d+)\)\s*(.+)', line)
        if premise_match:
            premise = premise_match.group(2).strip()
            # Only add premises that don't start with # (uncommented ones)
            if not line.startswith('#'):
                premises.append(premise)
                # Extract variables from premise
                vars_in_premise = re.findall(r'\b[A-Z]\b', premise)
                variables.update(vars_in_premise)
            continue
        
        # Find conclusion
        conclusion_match = re.match(r'Conclusion:\s*(.+)', line)
        if conclusion_match:
            conclusion = conclusion_match.group(1).strip()
            # Extract variables from conclusion
            vars_in_conclusion = re.findall(r'\b[A-Z]\b', conclusion)
            variables.update(vars_in_conclusion)
    
    return premises, conclusion, sorted(list(variables))

def generate_truth_table(input_filename):
    """Generate the truth table from input file"""
    # Check if input file exists
    if not os.path.exists(input_filename):
        print(f"Error: Input file '{input_filename}' not found.")
        return
    
    # Read and parse input file
    try:
        premises, conclusion, variables = read_input_file(input_filename)
    except Exception as e:
        print(f"Error reading input file: {e}")
        return
    
    if not premises:
        print("No premises found in input file.")
        return
    
    if not conclusion:
        print("No conclusion found in input file.")
        return
    
    # Calculate statistics
    num_vars = len(variables)
    num_premises = len(premises)
    num_rows = 2 ** num_vars
    
    print(f"Found {num_premises} premises and 1 conclusion")
    print(f"Variables: {variables}")
    
    # Generate output filenames
    base_name = os.path.splitext(input_filename)[0]
    csv_output_filename = f"{base_name}.out.csv"
    excel_output_filename = f"{base_name}.out.xlsx"
    
    # Create premise evaluators
    premise_evaluators = []
    for i, premise in enumerate(premises):
        try:
            evaluator = parse_logical_expression(premise, variables)
            premise_evaluators.append(evaluator)
            print(f"Premise {i+1}: {premise}")
        except Exception as e:
            print(f"Error parsing premise {i+1}: {premise}")
            print(f"Error: {e}")
            return
    
    # Create conclusion evaluator
    try:
        conclusion_evaluator = parse_logical_expression(conclusion, variables)
        print(f"Conclusion: {conclusion}")
    except Exception as e:
        print(f"Error parsing conclusion: {conclusion}")
        print(f"Error: {e}")
        return
    
    # Create headers - add index column first, then add numbers before each premise formula
    premise_cols = [f"{i+1}. {premise}" for i, premise in enumerate(premises)]
    header = ['Index'] + variables + premise_cols + ['All_Premises', conclusion, f'{conclusion} (Always)']
    
    # Generate all possible combinations
    rows = []
    
    for i, combination in enumerate(itertools.product([0, 1], repeat=num_vars), 1):
        # Create variable assignment
        var_dict = dict(zip(variables, combination))
        
        # Evaluate all premises
        premise_values = []
        for evaluator in premise_evaluators:
            try:
                result = evaluator(**var_dict)
                premise_values.append(int(result))
            except Exception as e:
                print(f"Error evaluating premise with {var_dict}: {e}")
                premise_values.append(0)
        
        # Check if all premises are true
        all_premises = all(premise_values)
        
        # Always evaluate conclusion for debugging
        try:
            conclusion_result = conclusion_evaluator(**var_dict)
            conclusion_value = int(conclusion_result)
        except Exception as e:
            print(f"Error evaluating conclusion with {var_dict}: {e}")
            conclusion_value = 0
        
        # For display, show conclusion value only when all premises are true
        if all_premises:
            conclusion_str = str(conclusion_value)
        else:
            conclusion_str = "-"
        
        # Create row - add index as first column, then variables, premises, etc.
        row = [i] + list(combination) + premise_values + [int(all_premises), conclusion_str, str(conclusion_value)]
        rows.append(row)
    
    # Find rows where all premises are true and count conclusions = 1
    premise_true_rows = []
    for row in rows:
        all_premises_col_index = len(variables) + 1 + len(premises)  # Index + Variables + Premises
        if row[all_premises_col_index] == 1:  # All_Premises column is 1
            premise_true_rows.append(row)
    
    print(f"\nRows where all premises are TRUE: {len(premise_true_rows)}")
    
    # Count how many of these rows have conclusion = 1
    if len(premise_true_rows) == 0:
        argument_valid = 0
        print("No rows where all premises are true")
    else:
        # Count rows where conclusion is 1
        conclusion_always_col_index = len(variables) + 1 + len(premises) + 2  # Last column
        conclusion_true_count = 0
        
        for row in premise_true_rows:
            row_index = row[0]
            conclusion_value = int(row[conclusion_always_col_index])  # Always column
            var_assignment = {var: row[i+1] for i, var in enumerate(variables)}
            print(f"  Row {row_index}: {var_assignment} -> Conclusion: {conclusion_value}")
            
            if conclusion_value == 1:
                conclusion_true_count += 1
        
        # The argument validity is the COUNT of rows where conclusion is true
        argument_valid = conclusion_true_count
        print(f"Rows with conclusion = 1: {conclusion_true_count}")
    
    print(f"FINAL: Argument validity: {argument_valid}")
    
    # Keep valid_rows for compatibility with rest of code
    valid_rows = premise_true_rows
    
    # Count 0s and 1s for each variable when All_Premises=1
    var_counts_0 = {}
    var_counts_1 = {}
    
    for var in variables:
        var_index = variables.index(var) + 1  # +1 because of Index column
        if valid_rows:
            var_counts_0[var] = sum(1 for row in valid_rows if row[var_index] == 0)
            var_counts_1[var] = sum(1 for row in valid_rows if row[var_index] == 1)
        else:
            var_counts_0[var] = "-"
            var_counts_1[var] = "-"
    
    # Create the summary row showing argument validity
    summary_row = ["-"] + ["-"] * len(variables)  # Index + variable columns
    summary_row += ["-"] * len(premises)  # Fill premise columns with "-"
    summary_row += ["-"]  # All_Premises column
    summary_row += [str(argument_valid)]  # Argument validity (count of valid conclusions)
    summary_row += [str(argument_valid)]  # Also add to the always-evaluated column
    
    # Create count rows for 0s and 1s
    count_0_row = ["0"] + ["-"] * len(variables)  # Index=0 + variable columns
    count_1_row = ["1"] + ["-"] * len(variables)  # Index=1 + variable columns
    
    # Fill in the counts for each variable
    for i, var in enumerate(variables):
        var_index = i + 1  # +1 because of Index column
        count_0_row[var_index] = str(var_counts_0[var])
        count_1_row[var_index] = str(var_counts_1[var])
    
    # Fill rest of count rows with "-"
    count_0_row += ["-"] * len(premises)  # Fill premise columns with "-"
    count_0_row += ["-"]  # All_Premises column
    count_0_row += ["-"]  # Conclusion columns
    count_0_row += ["-"]
    
    count_1_row += ["-"] * len(premises)  # Fill premise columns with "-"
    count_1_row += ["-"]  # All_Premises column
    count_1_row += ["-"]  # Conclusion columns
    count_1_row += ["-"]
    
    # Write to CSV file with emoji formatting (for compatibility)
    write_csv_file(csv_output_filename, header, rows, summary_row, count_0_row, count_1_row, 
                   num_vars, num_premises, num_rows, valid_rows, argument_valid, variables, premises)
    
    # Write to Excel file with actual color formatting (if openpyxl is available)
    if EXCEL_AVAILABLE:
        write_excel_file(excel_output_filename, header, rows, summary_row, count_0_row, count_1_row,
                        num_vars, num_premises, num_rows, valid_rows, argument_valid, variables, premises)
        print(f"\nTruth table with colors saved to: {excel_output_filename}")
    
    print(f"\nTruth table (CSV) saved to: {csv_output_filename}")

def write_csv_file(output_filename, header, rows, summary_row, count_0_row, count_1_row,
                   num_vars, num_premises, num_rows, valid_rows, argument_valid, variables, premises):
    """Write CSV file with emoji formatting"""
    try:
        with open(output_filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile, quoting=csv.QUOTE_ALL)
            writer.writerow(header)
            
            # Write data rows with emoji formatting for premise columns
            for row in rows:
                formatted_row = []
                for col_idx, value in enumerate(row):
                    # Check if this is a premise column (after variables but before All_Premises)
                    premise_start_idx = 1 + len(variables)  # After Index + Variables
                    premise_end_idx = premise_start_idx + len(premises)  # Before All_Premises
                    
                    is_premise_col = premise_start_idx <= col_idx < premise_end_idx
                    
                    if is_premise_col and value in [0, 1]:
                        if value == 0:
                            formatted_row.append(f"🎀{value}")  # Pink emoji for 0
                        else:
                            formatted_row.append(f"✅{value}")  # Green checkmark for 1
                    else:
                        formatted_row.append(str(value))
                
                writer.writerow(formatted_row)
            
            writer.writerow(summary_row)
            writer.writerow(count_0_row)
            writer.writerow(count_1_row)
            # Add statistics
            writer.writerow([f"# of variables: {num_vars}"])
            writer.writerow([f"# of Premises: {num_premises}"])
            writer.writerow([f"Number of rows in the truth table: {num_rows}"])
            if valid_rows:
                writer.writerow([f" Rows where all premises are true: {len(valid_rows)}"])
                writer.writerow([f" Conclusion values for these rows:"])
                for row in valid_rows:
                    row_index = row[0]
                    var_values = dict(zip(variables, row[1:num_vars+1]))
                    conclusion_val = row[-2] if row[-2] != "-" else row[-1]
                    writer.writerow([f"  Row {row_index}: {var_values} -> Conclusion: {conclusion_val}"])
                writer.writerow([f" Argument validity: {argument_valid} (Count of rows where conclusion is true)"])
            else:
                writer.writerow([" No rows where all premises are true."])
                writer.writerow([f" Argument validity: {argument_valid} (No valid premise combinations)"])

    except PermissionError:
        print(f"\nError: Cannot write to '{output_filename}' - file may be open in another program.")
        handle_permission_error(output_filename, header, rows, summary_row, count_0_row, count_1_row,
                               num_vars, num_premises, num_rows, valid_rows, argument_valid, variables, premises)

def write_excel_file(output_filename, header, rows, summary_row, count_0_row, count_1_row,
                     num_vars, num_premises, num_rows, valid_rows, argument_valid, variables, premises):
    """Write Excel file with actual color formatting"""
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Truth Table"
        
        # Define fill colors
        pink_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light pink
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        
        # Write header
        for col_idx, header_val in enumerate(header, 1):
            ws.cell(row=1, column=col_idx, value=header_val)
        
        # Write data rows with color formatting
        for row_idx, row in enumerate(rows, 2):  # Start from row 2 (after header)
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Check if this is a premise column (after variables but before All_Premises)
                premise_start_idx = 1 + len(variables) + 1  # After Index + Variables (1-indexed)
                premise_end_idx = premise_start_idx + len(premises)  # Before All_Premises
                
                is_premise_col = premise_start_idx <= col_idx < premise_end_idx
                
                if is_premise_col and value in [0, 1]:
                    if value == 0:
                        cell.fill = pink_fill
                    else:
                        cell.fill = green_fill
        
        # Write summary rows
        current_row = len(rows) + 2
        
        # Summary row
        for col_idx, value in enumerate(summary_row, 1):
            ws.cell(row=current_row, column=col_idx, value=value)
        current_row += 1
        
        # Count rows
        for col_idx, value in enumerate(count_0_row, 1):
            ws.cell(row=current_row, column=col_idx, value=value)
        current_row += 1
        
        for col_idx, value in enumerate(count_1_row, 1):
            ws.cell(row=current_row, column=col_idx, value=value)
        current_row += 1
        
        # Add statistics
        ws.cell(row=current_row, column=1, value=f"# of variables: {num_vars}")
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"# of Premises: {num_premises}")
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"Number of rows in the truth table: {num_rows}")
        current_row += 1
        
        if valid_rows:
            ws.cell(row=current_row, column=1, value=f" Rows where all premises are true: {len(valid_rows)}")
            current_row += 1
            ws.cell(row=current_row, column=1, value=" Conclusion values for these rows:")
            current_row += 1
            for row in valid_rows:
                row_index = row[0]
                var_values = dict(zip(variables, row[1:num_vars+1]))
                conclusion_val = row[-2] if row[-2] != "-" else row[-1]
                ws.cell(row=current_row, column=1, value=f"  Row {row_index}: {var_values} -> Conclusion: {conclusion_val}")
                current_row += 1
            ws.cell(row=current_row, column=1, value=f" Argument validity: {argument_valid} (Count of rows where conclusion is true)")
        else:
            ws.cell(row=current_row, column=1, value=" No rows where all premises are true.")
            current_row += 1
            ws.cell(row=current_row, column=1, value=f" Argument validity: {argument_valid} (No valid premise combinations)")
        
        # Save the workbook
        wb.save(output_filename)
        
    except PermissionError:
        print(f"\nError: Cannot write to Excel file '{output_filename}' - file may be open in another program.")
        print("Please close the file in Excel and try again.")
    except Exception as e:
        print(f"Error creating Excel file: {e}")

def handle_permission_error(output_filename, header, rows, summary_row, count_0_row, count_1_row,
                           num_vars, num_premises, num_rows, valid_rows, argument_valid, variables, premises):
    
    # Print statistics immediately after the CSV file is written
    print(f"# of variables: {num_vars}")
    print(f"# of Premises: {num_premises}")
    print(f"Number of rows in the truth table: {num_rows}")
    
    # Print summary
    if valid_rows:
        print(f"\nRows where all premises are true: {len(valid_rows)}")
        print("Conclusion values for these rows:")
        for row in valid_rows:
            row_index = row[0]  # Get the actual index from the first column
            var_values = dict(zip(variables, row[1:num_vars+1]))  # Skip index column
            conclusion_val = row[-2] if row[-2] != "-" else row[-1]
            print(f"  Row {row_index}: {var_values} -> Conclusion: {conclusion_val}")
        
        print(f"\nArgument validity: {argument_valid} (Count of rows where conclusion is true)")
    else:
        print("\nNo rows where all premises are true.")
        print(f"Argument validity: {argument_valid} (No valid premise combinations)")

def test_your_input():
    """Test function for your specific input"""
    variables = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    
    # Active premises (not commented out)
    premises = [
        "(A⊕B) → C",           # 1
        "(E ∨ F) ↔ G",         # 3  
        "(G ∧ H) → (A ∨ B)",   # 5
        "(¬E ∧ F) ↔ (C ∧ D)",  # 11
        "(A ∨ C) → (B ∨ ¬F)"   # 12
    ]
    
    conclusion = "(G ∧ H) ∧ (A ∨ B) ∧ ¬E"
    
    print("Testing your specific input:")
    print(f"Variables: {variables}")
    print("Active premises:")
    for i, premise in enumerate(premises, 1):
        print(f"  {i}: {premise}")
    print(f"Conclusion: {conclusion}")
    
    # Create evaluators
    premise_evaluators = []
    for i, premise in enumerate(premises):
        try:
            evaluator = parse_logical_expression(premise, variables)
            premise_evaluators.append(evaluator)
        except Exception as e:
            print(f"Error parsing premise {i+1}: {e}")
            return
    
    try:
        conclusion_evaluator = parse_logical_expression(conclusion, variables)
    except Exception as e:
        print(f"Error parsing conclusion: {e}")
        return
    
    # Test a few combinations manually
    print("\nTesting some specific combinations:")
    
    # Test case where we might expect premises to be true
    test_cases = [
        {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'E': 0, 'F': 0, 'G': 0, 'H': 0},
        {'A': 1, 'B': 0, 'C': 1, 'D': 1, 'E': 0, 'F': 1, 'G': 1, 'H': 1},
        {'A': 0, 'B': 1, 'C': 1, 'D': 1, 'E': 0, 'F': 1, 'G': 1, 'H': 1},
    ]
    
    for i, test_case in enumerate(test_cases):
        print(f"\nTest case {i+1}: {test_case}")
        
        premise_results = []
        for j, evaluator in enumerate(premise_evaluators):
            result = evaluator(**test_case)
            premise_results.append(result)
            print(f"  Premise {j+1}: {result}")
        
        all_premises = all(premise_results)
        conclusion_result = conclusion_evaluator(**test_case)
        
        print(f"  All premises true: {all_premises}")
        print(f"  Conclusion: {conclusion_result}")
        
        if all_premises:
            print(f"  -> This is a valid row! Conclusion should be true for valid argument")

def test_expression():
    """Test function for debugging specific expressions"""
    # First test the original case
    variables = ['B', 'C', 'D', 'G']
    expr = "(B⊕D)∧¬(C∧G)"
    
    evaluator = parse_logical_expression(expr, variables)
    
    # Test case from Row 2
    test_vars = {'B': 1, 'C': 0, 'D': 0, 'G': 0}
    result = evaluator(**test_vars)
    
    print(f"\nOriginal test case:")
    print(f"Expression: {expr}")
    print(f"Variables: {test_vars}")
    print(f"Result: {result} ({int(result)})")
    
    # Manual calculation
    B, C, D, G = 1, 0, 0, 0
    xor_BD = B != D  # 1 != 0 = True
    and_CG = C and G  # 0 and 0 = False
    not_and_CG = not and_CG  # not False = True
    final = xor_BD and not_and_CG  # True and True = True
    
    print(f"\nManual calculation:")
    print(f"B⊕D = {B}⊕{D} = {xor_BD}")
    print(f"C∧G = {C}∧{G} = {and_CG}")
    print(f"¬(C∧G) = ¬{and_CG} = {not_and_CG}")
    print(f"(B⊕D)∧¬(C∧G) = {xor_BD}∧{not_and_CG} = {final}")
    
    print("\n" + "="*50)
    test_your_input()

def main():
    """Main function to handle command line arguments"""
    if len(sys.argv) == 2 and sys.argv[1] == "test":
        test_expression()
        return
    
    if len(sys.argv) != 2:
        print("Usage: python script.py <input_filename>")
        print("       python script.py test  (for testing)")
        print("Example: python script.py problem.txt")
        print("\nInput file format:")
        print("1) premise1")
        print("2) premise2")
        print("...")
        print("Conclusion: conclusion_expression")
        sys.exit(1)
    
    input_filename = sys.argv[1]
    generate_truth_table(input_filename)

if __name__ == "__main__":
    main()